import os
import pandas as pd
import openai
import time
import json
from json import JSONDecodeError
from tqdm import tqdm
from OCR_Reader import CVsReader
import google.generativeai as genai
import textwrap
from openpyxl import Workbook
import shutil
import streamlit as st

# Define the path to the output CSV file and Excel file
output_csv_file_path = 'Output/CVs_Info_Extracted.csv'
output_excel_file_path = 'Output/CVs_Info_Extracted.xlsx'

class CVsInfoExtractor:
    def __init__(self, cvs_df):
        self.cvs_df = cvs_df
        
        with open('Engineered_Prompt/Prompt_New.txt', 'r') as file:
            self.prompt = file.read()
        
        openai.api_key = st.secrets["openai_secret_key"]

    def _call_gpt_for_cv_info_extraction(self, prompt, cv_content, model, temperature=0):
        completion_params = {
            'model': model,
            'messages': [{"role": "system", "content": prompt},
                         {"role": "user", "content": cv_content}],
            'temperature': temperature
        }
        response = openai.ChatCompletion.create(**completion_params)
        if 'choices' in response and len(response.choices) > 0:
            cleaned_response = response['choices'][0]['message']['content']
            try:
                json_response = json.loads(cleaned_response)
            except JSONDecodeError:
                json_response = None  
        else:
            json_response = None

        return json_response

    def _normalize_gpt_json_response(self, CV_Filename, json_response):
        CV_Filename_df = pd.DataFrame([CV_Filename], columns=['CV_Filename'])
        df_CV_Info_Json_normalized = pd.json_normalize(json_response)
        df = pd.concat([CV_Filename_df, df_CV_Info_Json_normalized], axis=1)
        return df

    def _write_response_to_file(self, df):
        if os.path.isfile(output_csv_file_path):
            df.to_csv(output_csv_file_path, mode='a', index=False, header=False)
        else:
            df.to_csv(output_csv_file_path, mode='w', index=False)

    def _gpt_pipeline(self, row, model='gpt-3.5-turbo'):
        CV_Filename = row['CV_Filename']
        CV_Content = row['CV_Content']
        time.sleep(5)
        json_response = self._call_gpt_for_cv_info_extraction(prompt=self.prompt, cv_content=CV_Content, model=model)
        df = self._normalize_gpt_json_response(CV_Filename, json_response)
        self._write_response_to_file(df)
        return json_response

    def _write_final_results_to_excel(self):
        df_to_excel = pd.read_csv(output_csv_file_path)
        df_to_excel.to_excel(output_excel_file_path)
        return df_to_excel

    def extract_cv_info(self):
        tqdm.pandas(desc="Progress")
        self.cvs_df['CV_Info_Json'] = self.cvs_df.progress_apply(self._gpt_pipeline, axis=1)
        st.write('CV Extraction Completed!')
        st.write('Saving Results to Output/CVs_Info_Extracted.xlsx')
        final_df = self._write_final_results_to_excel()
        return final_df
    
    def move_processed_cvs(self, source_directory):
        output_directory = "CV"  

        # Set to keep track of copied subfolders
        copied_subfolders = set()

        # Iterate through the DataFrame of extracted CVs
        for index, row in self.cvs_df.iterrows():
            cv_filename = row['CV_Filename']  # Get the original CV filename
            cv_file_path = os.path.join(source_directory, cv_filename)  # Full path of the original file
            cv_subfolder = os.path.basename(os.path.dirname(cv_file_path))  # Get the name of the parent folder
            
            # Define the source subfolder path
            source_subfolder_path = os.path.dirname(cv_file_path)

            # Create the destination folder for the subfolder if it doesn't exist
            destination_folder = os.path.join(output_directory, cv_subfolder)
            
            # Check if the subfolder has already been copied
            if cv_subfolder not in copied_subfolders:
                if not os.path.exists(destination_folder):
                    shutil.copytree(source_subfolder_path, destination_folder)
                    copied_subfolders.add(cv_subfolder)  # Mark this subfolder as copied
                    st.success(f"Copied folder {source_subfolder_path} to {destination_folder}")

            # Move the CV file to the destination folder
            destination_file_path = os.path.join(destination_folder, os.path.basename(cv_filename))
            shutil.move(cv_file_path, destination_file_path)
            st.success(f"Moved {cv_filename} to {destination_folder}")

    def fetch_experience_column(self, requires):
        try:
            df = pd.read_excel(output_excel_file_path)
            
            if "Experience" not in df.columns:
                return [], "Column 'Experience' not found in the header."
            
            responses = []
            for index, row in df.iterrows():
                experience = row["Experience"] if pd.notna(row["Experience"]) else ""
                response = self.google_api(experience, requires)
                responses.append(response)

            # Include the headers in the output
            return [df.columns.tolist()] + df.values.tolist(), responses

        except Exception as e:
            return [], f"An error occurred: {str(e)}"


class CVExtractorApp:
    def __init__(self):
        # Streamlit UI components
        st.title("CV Extractor")

        # CVs Directory Input
        self.cvs_directory_path = st.text_input("CVs Directory:")

        # Start Extraction Button
        if st.button("Start Extraction"):
            self.start_extraction()

        # Job Requirements Input
        self.requires_text = st.text_area("Job Requirements", height=150)

        # Fetch and Process Button
        if st.button("Fetch and Process Data"):
            self.fetch_and_process_data()

    def google_api(self, excel_docs, requires):
        genai.configure(api_key='AIzaSyDnpnOqVexPevNrZqMss9oqmaniiv6gGF0') 
        model = genai.GenerativeModel("gemini-pro")
        prompt = textwrap.dedent(f"""
            Based on the provided document, which is the experience of a job candidate. 
            And based on the requirement in job description. 
            Check whether the candidate's experience meets the requirement or not. 
            Split the requirement into single sentence, and evaluate each sentence. 
            The overall result will follow the following condition:
                - If the candidate experience meets 100% of the job description, return PASS. 
                - If the candidate experience meets 70% of the job description, return CONSIDER. 
                - If the candidate experience does not meet the job description, return FAIL. 
            Here is the sample format, please reply in this format: 
            "put the first sentence of the requirement here": "put its result here (PASS/FAIL)",
            ...
            "Overall result": "put overall result here (PASS/CONSIDER/FAIL)" 
                                
            Note that only return one status at a time (PASS or FAIL or CONSIDER), and don't return any other information. 
            The following is the provided document: 
            '{excel_docs}'. 
            And the following is the requirements:
            '{requires}'.
        """)
        answer = model.generate_content(prompt)
        return answer.candidates[0].content.parts[0].text

    def fetch_experience_column(self, requires):
        try:
            df = pd.read_excel(output_excel_file_path)
            
            if "Experience" not in df.columns:
                return [], "Column 'Experience' not found in the header."
            
            responses = []
            for index, row in df.iterrows():
                experience = row["Experience"] if pd.notna(row["Experience"]) else ""
                response = self.google_api(experience, requires)
                responses.append(response)

            return [df.columns.tolist()] + df.values.tolist(), responses

        except Exception as e:
            return [], f"An error occurred: {str(e)}"
        
    def write_to_excel(self, values, responses):
        wb = Workbook()
        ws = wb.active
        
        for col_index, header in enumerate(values[0], start=1):
            ws.cell(row=1, column=col_index, value=header)
        
        for row_index, row in enumerate(values[1:], start=2):
            for col_index, cell in enumerate(row, start=1):
                ws.cell(row=row_index, column=col_index, value=cell)

        result_col_index = len(values[0]) + 1
        ws.cell(row=1, column=result_col_index, value="Result")
        
        for row_index, response in enumerate(responses, start=2):
            ws.cell(row=row_index, column=result_col_index, value=response)
        
        os.makedirs('Output', exist_ok=True) 
        wb.save('Output/CV_filter_result.xlsx')

    def start_extraction(self):
        cvs_directory_path = self.cvs_directory_path

        if not os.path.exists(cvs_directory_path):
            st.error("The specified directory does not exist.")
            return

        # Log the start of the extraction
        st.info("Starting extraction...")

        # Create an instance of CVsReader
        cvs_reader = CVsReader(cvs_directory_path=cvs_directory_path)

        # Read CVs and assign to instance variable
        try:
            self.cvs_df = cvs_reader.read_cv()  
        except Exception as e:
            st.error(f"Error reading CVs: {str(e)}")
            return

        # Create an instance of CVsInfoExtractor
        cvs_info_extractor = CVsInfoExtractor(cvs_df=self.cvs_df)

        # Extract CV information
        try:
            extract_cv_info_dfs = cvs_info_extractor.extract_cv_info()
            st.success("Extraction completed successfully!")

            # Move processed CVs to the designated folder
            cvs_info_extractor.move_processed_cvs(cvs_directory_path)

        except Exception as e:
            st.error(f"Error during extraction: {str(e)}")


    def fetch_and_process_data(self):
        requires = self.requires_text.strip()

        if not requires:
            st.error("Job requirements are required.")
            return

        # Call fetch_experience_column with the job requirements
        values, responses = self.fetch_experience_column(requires)
        if values:
            self.write_to_excel(values, responses)
            st.success("CV Filtering completed. Results saved to 'Output/CV_filter_result.xlsx'.")
        else:
            st.error("Error: " + responses)


if __name__ == '__main__':
    app = CVExtractorApp()
