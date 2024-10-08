import os
import pandas as pd
import openai
import time
import json
from json import JSONDecodeError
from tqdm import tqdm
from OCR_Reader import CVsReader
import google.generativeai as genai
import textwrap
from openpyxl import Workbook
import shutil
import streamlit as st
import shutil

# Define the path to the output CSV file and Excel file
output_csv_file_path = 'Output/CVs_Info_Extracted.csv'
output_excel_file_path = 'Output/CVs_Info_Extracted.xlsx'

class CVsInfoExtractor:
    def __init__(self, cvs_df):
        self.cvs_df = cvs_df
        
        with open('Engineered_Prompt/Prompt_New.txt', 'r') as file:
            self.prompt = file.read()
        
        openai.api_key = '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'

    def _call_gpt_for_cv_info_extraction(self, prompt, cv_content, model, temperature=0):
        completion_params = {
            'model': model,
            'messages': [{"role": "system", "content": prompt},
                         {"role": "user", "content": cv_content}],
            'temperature': temperature
        }
        response = openai.ChatCompletion.create(**completion_params)
        if 'choices' in response and len(response.choices) > 0:
            cleaned_response = response['choices'][0]['message']['content']
            try:
                json_response = json.loads(cleaned_response)
            except JSONDecodeError:
                json_response = None  
        else:
            json_response = None

        return json_response

    def _normalize_gpt_json_response(self, CV_Filename, json_response):
        CV_Filename_df = pd.DataFrame([CV_Filename], columns=['CV_Filename'])
        df_CV_Info_Json_normalized = pd.json_normalize(json_response)
        df = pd.concat([CV_Filename_df, df_CV_Info_Json_normalized], axis=1)
        return df

    def _write_response_to_file(self, df):
        if os.path.isfile(output_csv_file_path):
            df.to_csv(output_csv_file_path, mode='a', index=False, header=False)
        else:
            df.to_csv(output_csv_file_path, mode='w', index=False)

    def _gpt_pipeline(self, row, model='gpt-3.5-turbo'):
        CV_Filename = row['CV_Filename']
        CV_Content = row['CV_Content']
        time.sleep(5)
        json_response = self._call_gpt_for_cv_info_extraction(prompt=self.prompt, cv_content=CV_Content, model=model)
        df = self._normalize_gpt_json_response(CV_Filename, json_response)
        self._write_response_to_file(df)
        return json_response

    def _write_final_results_to_excel(self):
        df_to_excel = pd.read_csv(output_csv_file_path)
        df_to_excel.to_excel(output_excel_file_path)
        return df_to_excel

    def extract_cv_info(self):
        tqdm.pandas(desc="Progress")
        self.cvs_df['CV_Info_Json'] = self.cvs_df.progress_apply(self._gpt_pipeline, axis=1)
        st.write('CV Extraction Completed!')
        st.write('Saving Results to Output/CVs_Info_Extracted.xlsx')
        final_df = self._write_final_results_to_excel()
        return final_df


    def fetch_experience_column(self, requires):
        try:
            df = pd.read_excel(output_excel_file_path)
            
            if "Experience" not in df.columns:
                return [], "Column 'Experience' not found in the header."
            
            responses = []
            for index, row in df.iterrows():
                experience = row["Experience"] if pd.notna(row["Experience"]) else ""
                response = self.google_api(experience, requires)
                responses.append(response)

            # Include the headers in the output
            return [df.columns.tolist()] + df.values.tolist(), responses

        except Exception as e:
            return [], f"An error occurred: {str(e)}"


class CVExtractorApp:
    def __init__(self):
        # Streamlit UI components
        st.title("CV Extractor")

        # CVs File Upload
        self.cvs_files = st.file_uploader("Upload CVs (PDF files)", type=["pdf"], accept_multiple_files=True)

        # Start Extraction Button
        if st.button("Start Extraction"):
            self.start_extraction()

        # Job Requirements Input
        self.requires_text = st.text_area("Job Requirements", height=150)

        # Fetch and Process Button
        if st.button("Fetch and Process Data"):
            self.fetch_and_process_data()

    def google_api(self, excel_docs, requires):
        genai.configure(api_key='AIzaSyDnpnOqVexPevNrZqMss9oqmaniiv6gGF0') 
        model = genai.GenerativeModel("gemini-pro")
        prompt = textwrap.dedent(f"""
            Based on the provided document, which is the experience of a job candidate. \
            And based on the requirement in job description. \
            Check whether the candidate's experience meets the requirement or not. \
            Split the requirement into single sentence, and evaluate each sentence. \
            The overall result will follow the following condition: \
                - If the candidate experience meets 100% of the job description, return PASS. \
                - If the candidate experience meets 70% of the job description, return CONSIDER. \
                - If the candidate experience does not meet the job description, return FAIL. \
            Here is the sample format, please reply in this format: \
            "put the first sentence of the requirement here": "put its result here (PASS/FAIL)",
            ...
            "Overall result": "put overall result here (PASS/CONSIDER/FAIL)" 
                                
            Note that only return one status at a time (PASS or FAIL or CONSIDER), and don't return any other information. \
            The following is the provided document: \
            '{excel_docs}'. 
            And the following is the requirements: \
            '{requires}'.
        """)
        answer = model.generate_content(prompt)
        return answer.candidates[0].content.parts[0].text

    def fetch_experience_column(self, requires):
        try:
            df = pd.read_excel(output_excel_file_path)
            
            if "Experience" not in df.columns:
                return [], "Column 'Experience' not found in the header."
            
            responses = []
            for index, row in df.iterrows():
                experience = row["Experience"] if pd.notna(row["Experience"]) else ""
                response = self.google_api(experience, requires)
                responses.append(response)

            return [df.columns.tolist()] + df.values.tolist(), responses

        except Exception as e:
            return [], f"An error occurred: {str(e)}"
        
    def write_to_excel(self, values, responses):
        wb = Workbook()
        ws = wb.active
        
        for col_index, header in enumerate(values[0], start=1):
            ws.cell(row=1, column=col_index, value=header)
        
        for row_index, row in enumerate(values[1:], start=2):
            for col_index, cell in enumerate(row, start=1):
                ws.cell(row=row_index, column=col_index, value=cell)

        result_col_index = len(values[0]) + 1
        ws.cell(row=1, column=result_col_index, value="Result")
        
        for row_index, response in enumerate(responses, start=2):
            ws.cell(row=row_index, column=result_col_index, value=response)
        
        os.makedirs('Output', exist_ok=True) 
        wb.save('Output/CV_filter_result.xlsx')


    def move_processed_files(self, uploaded_files, destination_directory):
        os.makedirs(destination_directory, exist_ok=True)

        destination_path = destination_directory

        # Loop through each uploaded file
        for uploaded_file in uploaded_files:
            # Create a temporary location to store the uploaded file
            temp_file_path = f"processed_{uploaded_file.name}"
            
            # Write the uploaded file to a temporary location
            with open(temp_file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            # Debugging: Check if the temporary file was created
            if not os.path.exists(temp_file_path):
                st.error(f"Temporary file not created: {temp_file_path}")
                continue

            # Move the file
            try:
                shutil.move(temp_file_path, destination_path)
                st.success(f"Moved {uploaded_file.name} to {destination_path}")
            except Exception as e:
                st.error(f"Error moving file {uploaded_file.name}: {str(e)}")
                continue

            if not os.path.exists(destination_path):
                st.error(f"File not found at destination: {destination_path}")


    def start_extraction(self):
        if not self.cvs_files:
            st.error("Please upload at least one CV file.")
            return

        # Log the start of the extraction
        st.info("Starting extraction...")

        # Create an instance of CVsReader
        cvs_reader = CVsReader()

        # Read CVs from uploaded files
        try:
            self.cvs_df = cvs_reader.read_cv_from_files(self.cvs_files)  # Update this method in CVsReader
        except Exception as e:
            st.error(f"Error reading CVs: {str(e)}")
            return

        # Create an instance of CVsInfoExtractor
        cvs_info_extractor = CVsInfoExtractor(cvs_df=self.cvs_df)

        # Extract CV information
        try:
            extract_cv_info_dfs = cvs_info_extractor.extract_cv_info()
            st.success("Extraction completed successfully!")

            # Move processed CVs if necessary
            # cvs_info_extractor.move_processed_cvs(cvs_directory_path)
            self.move_processed_files(self.cvs_files, "CV")

        except Exception as e:
            st.error(f"Error during extraction: {str(e)}")


    def fetch_and_process_data(self):
        requires = self.requires_text.strip()

        if not requires:
            st.error("Job requirements are required.")
            return

        # Call fetch_experience_column with the job requirements
        values, responses = self.fetch_experience_column(requires)
        if values:
            self.write_to_excel(values, responses)
            st.success("CV Filtering completed. Results saved to 'Output/CV_filter_result.xlsx'.")
        else:
            st.error("Error: " + responses)


if __name__ == '__main__':
    app = CVExtractorApp()
